import tkinter as tk
from tkinter import filedialog, messagebox
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment

class ExcelInputApp:
    def __init__(self, master):
        self.master = master
        self.master.title("Excel Input Application")
        self.master.geometry("800x600")

        self.excel_file = None
        self.sheet_names = []
        self.selected_sheet = tk.StringVar()

        self.create_widgets()

    def create_widgets(self):
        # File selection
        tk.Button(self.master, text="Select Excel File", command=self.select_file).pack(pady=10)
        self.file_label = tk.Label(self.master, text="No file selected")
        self.file_label.pack()

        # Sheet selection
        self.sheet_menu = tk.OptionMenu(self.master, self.selected_sheet, "")
        self.sheet_menu.pack(pady=10)

        # Input field
        self.input_frame = tk.Frame(self.master)
        self.input_frame.pack(pady=10, padx=10, fill=tk.BOTH, expand=True)

        tk.Label(self.input_frame, text="Paste your formatted text here:").pack()
        self.text_input = tk.Text(self.input_frame, height=20, width=80)
        self.text_input.pack(fill=tk.BOTH, expand=True)

        # Save button
        tk.Button(self.master, text="Save to Excel", command=self.save_to_excel).pack(pady=10)

    def select_file(self):
        self.excel_file = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx")])
        if self.excel_file:
            self.file_label.config(text=f"Selected file: {self.excel_file}")
            self.load_sheet_names()

    def load_sheet_names(self):
        workbook = load_workbook(self.excel_file)
        self.sheet_names = workbook.sheetnames
        self.selected_sheet.set(self.sheet_names[0])
        self.sheet_menu['menu'].delete(0, 'end')
        for sheet in self.sheet_names:
            self.sheet_menu['menu'].add_command(label=sheet, command=tk._setit(self.selected_sheet, sheet))

    def parse_input(self, text):
        # Split the input text into lines
        lines = text.strip().split('\n')
        
        # Parse each line
        parsed_data = []
        for line in lines:
            if line.startswith('|') and line.endswith('|'):
                # Remove leading and trailing '|' and split by '|'
                cells = [cell.strip() for cell in line.strip('|').split('|')]
                # Replace "<br>" with newline character
                cells = [cell.replace("<br>", "\n") for cell in cells]
                parsed_data.append(cells)

        return parsed_data

    def set_column_width(self, sheet):
        for column in sheet.columns:
            column_letter = get_column_letter(column[0].column)
            sheet.column_dimensions[column_letter].width = 100

    def save_to_excel(self):
        if not self.excel_file:
            messagebox.showerror("Error", "Please select an Excel file first.")
            return

        input_text = self.text_input.get("1.0", tk.END)
        parsed_data = self.parse_input(input_text)

        if not parsed_data:
            messagebox.showerror("Error", "No valid data to save.")
            return

        workbook = load_workbook(self.excel_file)
        sheet = workbook[self.selected_sheet.get()]

        # Find the first empty row
        row = sheet.max_row + 1

        # Write data to the sheet
        for data_row in parsed_data:
            for col, value in enumerate(data_row, start=1):
                cell = sheet.cell(row=row, column=col, value=value)
                cell.alignment = Alignment(wrap_text=True, vertical='center')
            row += 1

        # Autofit row heights
        for row in sheet.iter_rows(min_row=sheet.max_row - len(parsed_data) + 1, max_row=sheet.max_row):
            sheet.row_dimensions[row[0].row].auto_size = True

        # Set column width to 100
        self.set_column_width(sheet)

        workbook.save(self.excel_file)
        messagebox.showinfo("Success", f"{len(parsed_data)} rows saved to Excel successfully!")

        # Clear input field
        self.text_input.delete("1.0", tk.END)

if __name__ == "__main__":
    root = tk.Tk()
    app = ExcelInputApp(root)
    root.mainloop()